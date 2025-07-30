import openpyxl
from openpyxl import load_workbook

work_book = load_workbook('D:\PyGUI\CH-5-Event-Interaction\excel\emp_info.xlsx')

emp_sheet = work_book['Emp1']
print(emp_sheet.cell(row=1, column=2).value)
print(emp_sheet.cell(row=3, column=2).value)
print(emp_sheet.cell(row=3, column=3).value)
print(emp_sheet.cell(row=3, column=4).value)
print(emp_sheet.cell(row=3, column=5).value)


