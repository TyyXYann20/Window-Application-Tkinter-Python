import tkinter as tk
import openpyxl
from openpyxl import load_workbook
from tkinter import messagebox


path = 'D:\PyGUI\CH-5-Event-Interaction\excel\emp_info.xlsx'
work_book = load_workbook(path)
emp_sheet = work_book['Sheet1']

window = tk.Tk()
window.geometry("400x180")
window.title("Login From Excel")

def on_login():
    email = entered_email.get().strip()
    password = entered_password.get().strip()
    
    
    for row in emp_sheet.iter_rows(min_row=2, values_only=True):
        email_excel = str(row[2].strip())
        password_excel = str(row[3])
        
        if email == email_excel and password == password_excel:
            messagebox.showinfo("Login", "Login Successful!")
            return
        messagebox.showerror("Login Failed!", "Incorrect Email or Password!")
    

tk.Label(window,
         text="Email:",
         font=(20)).grid(row=0, column=0, sticky="w", pady=12)
tk.Label(window,
         text="Password:",
         font=(20)).grid(row=1, column=0, sticky="w", pady=12)


entered_email = tk.Entry(window, width=25, font=(20))
entered_password = tk.Entry(window, width=25, font=(20), show="*")

entered_email.grid(row=0, column=1, padx=12)
entered_password.grid(row=1, column=1, padx=12)

btn_login = tk.Button(window, text="Login", font=(20), 
                      cursor="hand2", relief="groove",
                      command=on_login)

btn_login.grid(row=2, column=1, sticky="e", padx=12)

window.mainloop()