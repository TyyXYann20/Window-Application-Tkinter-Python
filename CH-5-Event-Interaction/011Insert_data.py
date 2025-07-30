import openpyxl
from openpyxl import load_workbook

work_book = load_workbook('D:\PyGUI\CH-5-Event-Interaction\excel\emp_info.xlsx')

emp_sheet = work_book['Emp1']

emp_sheet.cell(row=4, column=2, value=2)
emp_sheet.cell(row=4, column=3, value='John')
emp_sheet.cell(row=4, column=4, value='abc20@gmail.com')
emp_sheet.cell(row=4, column=5, value='admin123$')
work_book.save('D:\PyGUI\CH-5-Event-Interaction\excel\emp_info.xlsx')

