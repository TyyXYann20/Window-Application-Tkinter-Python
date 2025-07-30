import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook

window = tk.Tk()
window.title("Register Page")
window.geometry("350x200")

path = 'D:\PyGUI\CH-5-Event-Interaction\excel\emp_info.xlsx'
work_book = load_workbook(path)
emp_sheet = work_book['Sheet1']


def register():
    emp_id = id_entry.get()
    name = username_entry.get()
    email = email_entry.get().strip()
    password = password_entry.get().strip()
    
    next_row = emp_sheet.max_row + 1
    emp_sheet.cell(row=next_row, column=1, value=emp_id)
    emp_sheet.cell(row=next_row, column=2, value=name)
    emp_sheet.cell(row=next_row, column=3, value=email)
    emp_sheet.cell(row=next_row, column=4, value=password)
    work_book.save(path)
   
# ID
tk.Label(window, text="ID:").grid(row=0, column=0, sticky="e", padx=10, pady=5)
id_entry = tk.Entry(window, width=25)
id_entry.grid(row=0, column=1, pady=5)
# Username
tk.Label(window, text="Username:").grid(row=1, column=0, sticky="e", padx=10, pady=5)
username_entry = tk.Entry(window, width=25)
username_entry.grid(row=1, column=1, pady=5)

tk.Label(window, text="Email:").grid(row=2, column=0, sticky="e", padx=10, pady=5)
email_entry = tk.Entry(window, width=25)
email_entry.grid(row=2, column=1, pady=5)

# Password
tk.Label(window, text="Password:").grid(row=3, column=0, sticky="e", padx=10, pady=5)
password_entry = tk.Entry(window, width=25, show="*")
password_entry.grid(row=3, column=1, pady=5)

register_btn = tk.Button(window, text="Register", command=register, width=15)
register_btn.grid(row=4, columnspan=2, pady=15)

window.mainloop()