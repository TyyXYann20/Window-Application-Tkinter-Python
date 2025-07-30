import openpyxl
from openpyxl import load_workbook

work_book = load_workbook('D:\PyGUI\CH-5-Event-Interaction\excel\emp_info.xlsx')

emp_sheet = work_book['Sheet1']

next_row = emp_sheet.max_row + 1
emp_sheet.cell(row=next_row, column=1, value=3)
emp_sheet.cell(row=next_row, column=2, value='Sara')
emp_sheet.cell(row=next_row, column=3, value='sara25@gmail.com')
emp_sheet.cell(row=next_row, column=4, value='admin123$')
work_book.save('D:\PyGUI\CH-5-Event-Interaction\excel\emp_info.xlsx')

print(emp_sheet.max_row)

