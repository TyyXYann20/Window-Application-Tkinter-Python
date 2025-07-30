import openpyxl
from openpyxl import load_workbook

work_book = load_workbook('D:\PyGUI\CH-5-Event-Interaction\excel\sales_data.xlsx')
sales_sheet= work_book.active
details_sheet = work_book['Details']


# print(sales_sheet.cell(row=2, column=2).value)
# print(sales_sheet.cell(row=23, column=6).value)
for i in range(1, 11):
    print(details_sheet.cell(row=i, column=1).value)